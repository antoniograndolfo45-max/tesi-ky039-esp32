import threading, time, re, queue, sys, os, datetime
from collections import deque

import tkinter as tk   
from tkinter import ttk, messagebox, filedialog

import serial
import serial.tools.list_ports as lp

import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

import pandas as pd  # per export Excel

# ================== CONFIG BASE ==================
BAUD_PREDEFINITO = 115200
SUGGERIMENTI_PORTA_AUTO = ("usbmodem", "usbserial", "wch", "ch340")
PUNTI_MASSIMI = 300  # punti nel grafico
# Regex per il protocollo che usa il firmware Arduino
REGEX_RIGA_BPM = re.compile(r"^\s*BPM:\s*([0-9]+(?:\.[0-9]+)?)")
REGEX_METRICHE  = re.compile(r"METRICS\s+baseline=([0-9.]+)\s+peak=([0-9.]+)\s+dHR=([0-9.]+)\s+tpeak=([0-9.]+)\s+recov60=([0-9.]+)")

# ================== THREAD LETTURA SERIALE ==================
class SerialReader(threading.Thread):
    def __init__(self, porta, baud, coda_uscita, su_errore):
        super().__init__(daemon=True) 
        #il thread daemon non impedisce al programma di chiudersi- se chiudo l'interfaccai si chiude anche il lettore seriale
        self.porta = porta   # salvo sulla ISTANZA la porta seriale
        self.baud = baud  # salvo il baud rate
        self.coda_uscita = coda_uscita # coda thread-safe per comunicare con la GUI
        self.su_errore = su_errore # funzione/callback per mostrare errori
        self._ferma = False # flag interno per fermare il thread 
        self.seriale = None # qui salverai l’oggetto serial.Serial quando lo apri

    def ferma(self):        #serve per fermare la lettura seriale
        self._ferma = True
        try:
            if self.seriale and self.seriale.is_open:
                self.seriale.close()
        except Exception:
            pass

    def esegui(self):  #ciclo usato per leggere la porta seriale
        try:
            self.seriale = serial.Serial(self.porta, self.baud, timeout=1) #apre la porta seriale 
            time.sleep(2)  # e conta fino a due per il reset di ESP32
            self.coda_uscita.put(("STATUS", f"Connesso a {self.porta} @ {self.baud}")) #messaggio ceh verrà mostrato
        except Exception as e:
            self.su_errore(f"Errore apertura seriale: {e}") #messaggio se non si connette
            return

        buff = bytearray()
        while not self._ferma: #il ciclo gira finche stop vine messo true da stop()
            try:
                blocco = self.seriale.read(256) #legge fino a 256 byte del seriale,se non arriva nulla ricomincia altrimenti salva in buffer
                if not blocco:
                    continue
                buff.extend(blocco)
                while b"\n" in buff:
                    linea, _, buff = buff.partition(b"\n") #legge la linea, mette il separatore e lascia lo spazio libero dopo ancora da processare 
                    try:
                        s = linea.decode(errors="ignore").strip() # .decode trasforma i byte in stringa e .strip toglie gli spazi
                        self.coda_uscita.put(("LINE", s)) #imvia la riga all'interfacca in modo ripulito dagli spazi
                    except Exception:
                        pass
            except Exception as e:
                self.su_errore(f"Errore lettura seriale: {e}")
                break

    # Thread.start() chiama run(): lo manteniamo e deleghiamo a esegui()
    def run(self):
        self.esegui()

# ================== APP TKINTER ==================
class App(tk.Tk): #crea l'oggetto della finistra principale
    def __init__(self): #la inizializza con titolo,dimensione e colore
        super().__init__()
        self.title("Arduino Heart Monitor-v3")
        self.geometry("1000x700")
        self._bg_base = self.cget("bg")
        top = ttk.Frame(self); top.pack(fill="x", padx=10, pady=8) #Crea un contenitore orizzontale con margini
        ttk.Label(top, text="Porta:").pack(side="left")
        self.combo_porta = ttk.Combobox(top, width=30, state="readonly") #ComboBox per scegliere la porta seriale "readonly"
        self.combo_porta.pack(side="left", padx=6)
        ttk.Label(top, text="Baud:").pack(side="left", padx=(12,2))
        self.var_baud = tk.StringVar(value=str(BAUD_PREDEFINITO)) #Variabile Tk che contiene il baud attuale così la ComboBox può leggere/scrivere 
        self.combo_baud = ttk.Combobox(top, width=8, state="readonly",
                                       values=["9600","19200","38400","57600","115200"],
                                       textvariable=self.var_baud) #può scegliere tra i valori proposti
        self.combo_baud.pack(side="left")
        ttk.Button(top, text="Refresh", command=self.aggiorna_porte).pack(side="left", padx=(8,0))
        ttk.Button(top, text="Connetti", command=self.connetti).pack(side="left", padx=6) #chiama self.connect() che crea e avvia il thread 
        ttk.Button(top, text="Disconnetti", command=self.disconnetti).pack(side="left") #chiama self.disconnect() che ferma il thread e chiude porta
        self.stato = tk.StringVar(value="Non connesso")
        ttk.Label(top, textvariable=self.stato).pack(side="right") # etichetta che visualizza il valore del self status
        controlli = ttk.Frame(self); controlli.pack(fill="x", padx=10, pady=6)
        self.var_bpm = tk.StringVar(value="--") # Variabile Tk che conterrà il BPM corrente 
        ttk.Label(controlli, textvariable=self.var_bpm, font=("Helvetica", 40)).pack(side="left") #Etichetta grande (font 40 pt) che mostra i BPM
        barra_strumenti = ttk.Frame(controlli); barra_strumenti.pack(side="right") #frame sulla destra che conterrà i pulsanti.
        ttk.Button(barra_strumenti, text="Protocollo 30–30–120",command=self.esegui_protocollo).pack(side="left", padx=6) #Avvia la sequenza 
        ttk.Button(barra_strumenti, text="Start (Baseline)",
                   command=lambda: self.invia_comando("CMD:START")).pack(side="left", padx=6) #Manda manualmente il comando START al device 
        ttk.Button(barra_strumenti, text="Stand",
                   command=lambda: self.invia_comando("CMD:STAND")).pack(side="left", padx=6) #Manda manualmente STAND 
        ttk.Button(barra_strumenti, text="Reset device",
                   command=lambda: self.invia_comando("CMD:RESET")).pack(side="left", padx=6) #Manda RESET per riportare la macchina a IDLE 
        ttk.Button(barra_strumenti, text="Esporta Excel…",
                   command=self.esporta_excel).pack(side="left", padx=6) #Apre il dialog “Salva con nome” ed esporta le metriche raccolte
        ttk.Button(barra_strumenti, text="Salva grafico PNG…",
                command=self.salva_png).pack(side="left", padx=6) #salva grafico
        # ---- pannello soglie/allerta ----
        tune = ttk.LabelFrame(self, text="Soglie e allerta"); tune.pack(fill="x", padx=10, pady=6) #Crea un riquadro con titolo “Soglie e allerta”
        self.soglia_dhr_min  = tk.DoubleVar(value=10.0)
        self.soglia_dhr_max  = tk.DoubleVar(value=30.0)
        self.soglia_picco_max = tk.DoubleVar(value=120.0)
        self.soglia_margine_recupero = tk.DoubleVar(value=10.0)

        for label, var in [("ΔHR min", self.soglia_dhr_min), ("ΔHR max", self.soglia_dhr_max),("Peak max", self.soglia_picco_max),("Recovmargin (+bpm)", self.soglia_margine_recupero)]:
            
            f = ttk.Frame(tune); f.pack(side="left", padx=8, pady=4)
            ttk.Label(f, text=label).pack()
            ttk.Entry(f, width=6, textvariable=var).pack() 
            #crea sezioni con casella di input per personalizzare le soglie

        # ---- metriche ----
        met = ttk.LabelFrame(self, text="Metriche ortostatico"); met.pack(fill="x", padx=10, pady=6) #crea riquadro con titolo
        self.m_basale = tk.StringVar(value="—")
        self.m_picco     = tk.StringVar(value="—")
        self.m_dhr      = tk.StringVar(value="—")
        self.m_tpicco    = tk.StringVar(value="—")
        self.m_recov    = tk.StringVar(value="—")
        self.var_interpretazione = tk.StringVar(value="Interpretazione: —") #spazi vuoti che la gui riempie quando arrivano dal seriale

        grid = ttk.Frame(met); grid.pack(fill="x", padx=6, pady=4)
        labels = [("Baseline", self.m_basale), ("Peak", self.m_picco),
                  ("ΔHR", self.m_dhr), ("t_peak (s)", self.m_tpicco),
                  ("Recov60", self.m_recov)] # il testo dell’etichetta e la StringVar che conterrà il valore da mostrare.
        for i, (lab, var) in enumerate(labels): #cicla su tuttte le coppie
            ttk.Label(grid, text=f"{lab}:").grid(row=0, column=2*i, sticky="w", padx=4) #crea etichetta nome e posizione geometrica
            ttk.Label(grid, textvariable=var).grid(row=0, column=2*i+1, sticky="w", padx=8) #crea etichetta del valore collegata a StringVar che si aggiorna e la mette accanto al nome
    
        ttk.Label(met, textvariable=self.var_interpretazione,
                  font=("Helvetica", 11, "italic"), foreground="blue").pack(anchor="w", padx=6, pady=4) #mostra l’interpretazione , che viene aggiornata dalla funzione di valutazione soglie quando ricevi le metriche 

        # ---- grafico ----
        self.figura = Figure(figsize=(8.7,3.8), dpi=100) #foglio del grafico
        self.assi  = self.figura.add_subplot(111) #aggiunge assi
        self.assi.set_xlabel("Campioni")
        self.assi.set_ylabel("BPM")
        self.linea, = self.assi.plot([], []) #sfruta gli elementi arrivati da plot()
        self.assi.grid(True)

        tela = FigureCanvasTkAgg(self.figura, master=self) #canvas Tkinter che contiene la figura Matplotlib
        tela.draw() #disegna figura
        tela.get_tk_widget().pack(fill="both", expand=True, padx=10, pady=6) #inserisce widget Tk del canvas dove both riempie in larghezza e altezza e expand ridimensiona quando riempi la finestra
        self.tela = tela #salva il canvas

        # ---- menù ----
        barra_menu = tk.Menu(self) # crea la barra di menu in alto con le relative voci
        menu_test = tk.Menu(barra_menu, tearoff=0)
        menu_test.add_command(label="Avvia protocollo 30–30–120", command=self.esegui_protocollo)
        menu_test.add_command(label="Reset (device)", command=lambda: self.invia_comando("CMD:RESET"))
        barra_menu.add_cascade(label="Test", menu=menu_test)

        menu_file = tk.Menu(barra_menu, tearoff=0)
        menu_file.add_command(label="Esporta Excel…", command=self.esporta_excel)
        menu_file.add_separator()
        menu_file.add_command(label="Esci", command=self.alla_chiusura)
        menu_file.add_command(label="Salva grafico PNG…", command=self.salva_png)
        menu_file.add_separator()
        barra_menu.add_cascade(label="File", menu=menu_file)
        self.config(menu=barra_menu) #Imposta la menubar appena creata come menu della finestra

        # ---- dati runtime ----
        self.serie_bpm = deque(maxlen=PUNTI_MASSIMI) #crea coda per gli ultimi BPM letti limitando la lunghezza e scartando i vecchi
        self.coda = queue.Queue() #Crea una coda thread-safe usata dal thread seriale per inviare messaggi alla GUI.
        self.lettore = None #è il thread che legge la porta e riomane none finche non clicci connetti
        self.righe = []  # dati per export Excel

        self.aggiorna_porte()
        self.after(100, self.leggi_coda)
        self.after(250, self.aggiorna_grafico)

    # ===== connessione =====
    def aggiorna_porte(self):
        porte = [p.device for p in lp.comports()] #chiede le porte seriali disponibili
        self.combo_porta["values"] = porte #riempie i valori con quelle porte seriali
        selezionata = None                                    #cerca di scegliere autonomamente una porta seriale confrontando
        for p in porte:
            minuscolo = p.lower()
            if any(h in minuscolo for h in SUGGERIMENTI_PORTA_AUTO):
                selezionata = p; break 
        self.combo_porta.set(selezionata or (porte[0] if porte else "")) #seleziona quella più probabile

    def su_errore(self, msg):
        self.stato.set("Errore")
        messagebox.showerror("Seriale", msg) #mostra messaggio errore

    def connetti(self):
        if self.lettore: #evita di connettersi due volte
            return
        porta = self.combo_porta.get().strip()
        if not porta:
            self.su_errore("Nessuna porta selezionata.")
            return #legge la porta scelta
        baud = int(self.var_baud.get()) #legge il baud
        self.lettore = SerialReader(porta, baud, self.coda, self.su_errore)
        self.lettore.start() #legge il seriale e lo avvia start()
        self.stato.set(f"Connessione a {porta}...") #aggiorna stato nella gui

    def disconnetti(self): 
        if self.lettore: #se c'è porta attiva chiude il seriale
            self.lettore.ferma()
            self.lettore = None
        self.stato.set("Disconnesso") #aggiorna lo stato a disconnesso

    def invia_comando(self, cmd): #invia comandi a ESP32
        if not self.lettore or not self.lettore.seriale or not self.lettore.seriale.is_open: #controlla se esiste lettore e se porta è aperta
            self.su_errore("Non connesso alla seriale.")
            return
        try:
            self.lettore.seriale.write((cmd + "\n").encode()) #scrive comando su seriale
        except Exception as e:
            self.su_errore(f"Invio comando fallito: {e}")

    # ===== protocollo ===== #serve per rannare il protocollo
    def esegui_protocollo(self):
        self.invia_comando("CMD:START")
        self.stato.set("Baseline in corso… (30 s)")
        self.after(30000, self._protocollo_stand)

    def _protocollo_stand(self):
        self.invia_comando("CMD:STAND")
        self.stato.set("Post-stand… (picco entro 30 s, recupero 60–120 s)")
        
    # =====  righe =====
    def leggi_coda(self):  #metodo per leggere la coda del seriale
        try:
            while True:
                typ, payload = self.coda.get_nowait() #legge
                if typ == "STATUS":
                    self.stato.set(payload) #aggiorna etichetta
                elif typ == "LINE":
                    self.gestisci_riga(payload) #se è una linea la passa a parser gestisci_riga()
        except queue.Empty:
            pass
        self.after(100, self.leggi_coda) #ripianifica se stessa

    def gestisci_riga(self, s): #interpreta i dati ricevuti da arduino e li salva per esportarli su excel(s è una stringa ricevuta dal seriale)
        m = REGEX_RIGA_BPM.match(s)
        if m:
            bpm = float(m.group(1)) #prende valore numero e lo converte in numero
            self.serie_bpm.append(bpm) #aggiunge a serie di valori per il grafico
            self.var_bpm.set(f"{bpm:.1f}") 
            return

        m2 = REGEX_METRICHE.search(s) #se la riga ricevuta è del tipo METRICS
        if m2:
            b, p, d, tp, r = map(float, m2.groups()) #converte tutto in variabili numeriche
            self.m_basale.set(f"{b:.1f}")
            self.m_picco.set(f"{p:.1f}")
            self.m_dhr.set(f"{d:.1f}")
            self.m_tpicco.set(f"{tp:.1f}")
            self.m_recov.set(f"{r:.1f}")
            # aggiorna etichette con valori calcolati
            self.valuta_e_avvisa(b, p, d, r, tp) #confronta con soglie
            # salva in memoria per Excel
            esito = "OK" if "OK " in self.stato.get() else "ATTENZIONE"
            self.righe.append({
                "timestamp": datetime.datetime.now().isoformat(timespec="seconds"),
                "baseline": round(b, 1),
                "peak": round(p, 1),
                "dHR": round(d, 1),
                "t_peak_s": round(tp, 1),
                "recov60": round(r, 1),
                "esito": esito,
                "interpretazione": self.var_interpretazione.get().replace("Interpretazione: ", "")
            })
            #aggiunge dati ad una lista che serve per esportare su excel

    # ===== valutazione soglie + interpretazione =====
    def valuta_e_avvisa(self, baseline, peak, dhr, recov60, tpeak):
        reasons = []

        # HR_baseline
        if baseline < 50:
            reasons.append("Bradicardia (<50 bpm)")
        elif baseline > 100:
            reasons.append("Tachicardia a riposo (>100 bpm)")

        # ΔHR
        if dhr < 10:
            reasons.append("ΔHR troppo basso (<10) → sospetta ipotensione/disautonomia")
        elif dhr > 40:
            reasons.append("ΔHR troppo alto (>40) → possibile POTS/ansia")

        # HR_peak
        if peak > 120:
            reasons.append("Peak >120 bpm → iper-adrenergico")

        # t_peak
        if tpeak > 30:
            reasons.append("t_peak >30s → reattività simpatica lenta")

        # recupero
        if (recov60 - baseline) > 15:
            reasons.append("Recupero lento (>15 bpm sopra baseline a 60–120s)")

        if not reasons:
            self.stato.set("Risultato: OK (entro soglie)")
            self.lampeggia("#b7f7c3")
            self.var_interpretazione.set("Interpretazione: Risposta normale")
        else:
            self.stato.set("Risultato: ATTENZIONE ")
            self.lampeggia("#ffb3b3")
            self.bell()
            self.var_interpretazione.set("Interpretazione: " + " | ".join(reasons))

   
    # ===== export Excel =====
    def esporta_excel(self):
        if not self.righe:
            messagebox.showinfo("Export", "Nessun dato da esportare.")
            return
        try:
            df = pd.DataFrame(self.righe, columns=[
                "timestamp","baseline","peak","dHR","t_peak_s","recov60","esito","interpretazione"
            ]) 
            #converte la lista di dizionari self.rows in un DataFrame pandas con le colonne nell’ordine desiderato.
            nomefile = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel file","*.xlsx")],
                initialfile=f"ortho_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            ) #apre la finestra "salva con nome" e scegli dove metterlo
            if not nomefile:
                return
           
            # Scrittura con formattazione usando openpyxl
            with pd.ExcelWriter(nomefile, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Risultati")
                ws = writer.sheets["Risultati"]
            # dataFrame nel foglio “Risultati” 
                # Stile intestazioni
                from openpyxl.styles import Font, PatternFill
                for cell in ws[1]:
                    cell.font = Font(bold=True)

                # Formattazione condizionale sulle righe
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    esito_cell = row[6]  # colonna "esito"
                    if esito_cell.value == "OK":
                        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # verde
                    else:
                        fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # rosso
                    for c in row:
                        c.fill = fill

            self.stato.set(f"Esportato: {nomefile}")
            messagebox.showinfo("Export", f"File salvato con formattazione:\n{nomefile}")
        except Exception as e:
            self.su_errore(f"Export fallito: {e}")
            
    # ===== grafico =====
    def salva_png(self, svg: bool = False):
        
        try:
            # Evita di salvare un grafico vuoto
            if not self.serie_bpm:
                messagebox.showinfo("Salva grafico", "Non ci sono dati nel grafico da salvare.")
                return

            est = ".svg" if svg else ".png"
            tipo_descr = "SVG" if svg else "PNG"
            nome_suggerito = f"grafico_bpm_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}{est}"

            path = filedialog.asksaveasfilename(
                defaultextension=est,
                filetypes=[("Immagine PNG","*.png"), ("Immagine SVG","*.svg")] if not svg
                         else [("Immagine SVG","*.svg"), ("Immagine PNG","*.png")],
                initialfile=nome_suggerito,
                title=f"Salva grafico {tipo_descr}"
            )
            if not path:
                return

            # Salvataggio (DPI più alto per PNG)
            if svg:
                self.figura.savefig(path, bbox_inches="tight")
            else:
                self.figura.savefig(path, dpi=200, bbox_inches="tight")

            self.stato.set(f"Grafico salvato: {path}")
            messagebox.showinfo("Salva grafico", f"Grafico salvato con successo:\n{path}")

        except Exception as e:
            self.su_errore(f"Salvataggio grafico fallito: {e}")
    def aggiorna_grafico(self):
        y = list(self.serie_bpm)
        x = list(range(len(y)))
        self.linea.set_data(x, y) #prende gli ultimi BPM memorizzati nella deque self.bpm_series  e costruisce un asse x corrispondente 
        if y:
            ymin = max(30, int(min(y)) - 5)
            ymax = min(200, int(max(y)) + 5)
            if ymin >= ymax: ymax = ymin + 5
            self.assi.set_ylim(ymin, ymax)
            self.assi.set_xlim(0, max(30, len(y)))
        self.tela.draw_idle()
        self.after(200, self.aggiorna_grafico) #ripianifica se stessa cosi si aggiorna in modo continuo

    def lampeggia(self, color, times=3, interval=180):
        def _step(n):
            self.configure(bg=(color if n % 2 == 0 else self._bg_base))
            if n < times*2:
                self.after(interval, _step, n+1)
            else:
                self.configure(bg=self._bg_base)
        _step(0) #fa lampeggiare lo sfondo della finestra

    def alla_chiusura(self):
        self.disconnetti()
        self.destroy() #chiude porta seriale

# ================== MAIN ==================
if __name__ == "__main__":
    app = App()
    app.protocol("WM_DELETE_WINDOW", app.alla_chiusura)
    app.mainloop()